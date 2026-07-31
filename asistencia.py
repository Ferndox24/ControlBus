import openpyxl
from datetime import datetime



def registrar_asistencia(trabajador, archivo, datos_qr=None):

    try:

        wb = openpyxl.load_workbook(
            archivo
        )

        ws = wb.active



        fila_trabajador = trabajador["FILA"]



        fila_encabezado = None
        columnas = {}



        # Buscar fila de encabezados

        for numero_fila in range(
            1,
            ws.max_row + 1
        ):


            valores = []


            for celda in ws[numero_fila]:


                if celda.value:

                    valores.append(
                        str(celda.value).strip().upper()
                    )



            if "ASISTENCIA" in valores:

                fila_encabezado = numero_fila
                break



        if fila_encabezado is None:

            print(
                "No se encontró encabezado"
            )

            return False





        # Obtener columnas existentes

        for columna, celda in enumerate(
            ws[fila_encabezado],
            start=1
        ):

            if celda.value:

                columnas[
                    str(celda.value).strip().upper()
                ] = columna





        # Crear columna SERIAL si no existe

        if "SERIAL" not in columnas:


            nueva_columna = ws.max_column + 1


            ws.cell(
                fila_encabezado,
                nueva_columna
            ).value = "SERIAL"


            columnas["SERIAL"] = nueva_columna





        # Crear columna MRZ si no existe

        if "MRZ" not in columnas:


            nueva_columna = ws.max_column + 1


            ws.cell(
                fila_encabezado,
                nueva_columna
            ).value = "MRZ"


            columnas["MRZ"] = nueva_columna





        columna_asistencia = columnas["ASISTENCIA"]





        # Revisar si ya está registrado

        asistencia_actual = ws.cell(
            fila_trabajador,
            columna_asistencia
        ).value



        if asistencia_actual:


            print(
                "Ya registrado:",
                asistencia_actual
            )


            return False





        fecha = datetime.now().strftime(
            "%d/%m/%Y %H:%M"
        )



        # Registrar asistencia

        ws.cell(
            fila_trabajador,
            columna_asistencia
        ).value = f"SI - {fecha}"





        # Guardar datos QR

        if datos_qr:


            ws.cell(
                fila_trabajador,
                columnas["SERIAL"]
            ).value = datos_qr.get(
                "SERIAL",
                ""
            )


            ws.cell(
                fila_trabajador,
                columnas["MRZ"]
            ).value = datos_qr.get(
                "MRZ",
                ""
            )





        wb.save(
            archivo
        )



        print(
            "Asistencia registrada correctamente"
        )


        return True





    except Exception as error:


        print(
            "Error:",
            error
        )


        return False