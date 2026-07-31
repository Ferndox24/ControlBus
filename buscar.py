
import openpyxl


def limpiar_rut(rut):

    if rut is None:
        return ""

    return (
        str(rut)
        .replace(".", "")
        .replace("-", "")
        .replace(" ", "")
        .replace("\n", "")
        .upper()
    )



def buscar_trabajador(rut, archivo):


    try:

        wb = openpyxl.load_workbook(
            archivo,
            data_only=True
        )


        ws = wb.active



        # Buscar fila de encabezados

        fila_encabezado = None


        for numero_fila in range(
            1,
            ws.max_row + 1
        ):


            valores = []


            for celda in ws[numero_fila]:

                if celda.value:

                    valores.append(
                        str(celda.value)
                        .strip()
                        .upper()
                    )


            if "RUT" in valores and "NOMBRE" in valores:

                fila_encabezado = numero_fila
                break



        if fila_encabezado is None:

            print("No se encontraron encabezados")

            return None




        encabezados = {}



        for columna, celda in enumerate(
            ws[fila_encabezado],
            start=1
        ):


            if celda.value:

                encabezados[
                    str(celda.value)
                    .strip()
                    .upper()
                ] = columna



        columna_rut = encabezados["RUT"]

        columna_nombre = encabezados["NOMBRE"]




        for fila in range(
            fila_encabezado + 1,
            ws.max_row + 1
        ):


            rut_excel = limpiar_rut(

                ws.cell(
                    fila,
                    columna_rut
                ).value

            )



            if rut_excel == limpiar_rut(rut):


                return {


                    "FILA": fila,


                    "RUT": ws.cell(
                        fila,
                        columna_rut
                    ).value,


                    "NOMBRE": ws.cell(
                        fila,
                        columna_nombre
                    ).value

                }




        return None



    except Exception as error:


        print(
            "Error buscando trabajador:",
            error
        )

        return None

